"""
PREDICT - Vehicle Intelligence Platform
Copyright © 2026 PREDICT
All rights reserved.

This file is proprietary and confidential.
Unauthorized copying, modification, distribution, or use is strictly prohibited.

Created: January 2026
Module: Data Management Tab

Data Management Tab
Centralized data management interface for:
- Export/Import data
- Backup management
- Data cleanup and optimization
- Storage usage monitoring
"""

import logging
import os
import shutil
import json
import csv
import zipfile
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, Dict, Any, List, Tuple

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QGroupBox, QFormLayout,
    QLineEdit, QComboBox, QProgressBar, QMessageBox, QFileDialog,
    QCheckBox, QSpinBox, QTabWidget, QScrollArea, QFrame
)
from PySide6.QtCore import Qt, Signal, QThread, QTimer
from PySide6.QtGui import QFont, QColor

# Import centralized configuration
try:
    from config import get_config
    CONFIG = get_config()
except ImportError:
    CONFIG = None

# Import data export and backup modules
try:
    from desktop_data_export import DataExportSystem
    DATA_EXPORT_AVAILABLE = True
except ImportError:
    DATA_EXPORT_AVAILABLE = False

try:
    from backup_manager import BackupManager
    BACKUP_MANAGER_AVAILABLE = True
except ImportError:
    BACKUP_MANAGER_AVAILABLE = False

logger = logging.getLogger(__name__)


class StorageAnalyzer(QThread):
    """Background thread for analyzing storage usage"""
    
    analysis_complete = Signal(dict)
    
    def __init__(self, data_dir: str):
        super().__init__()
        self.data_dir = data_dir
    
    def run(self):
        """Analyze storage usage"""
        try:
            total_size = 0
            file_count = 0
            folder_count = 0
            
            for root, dirs, files in os.walk(self.data_dir):
                folder_count += len(dirs)
                for file in files:
                    file_path = os.path.join(root, file)
                    try:
                        total_size += os.path.getsize(file_path)
                        file_count += 1
                    except OSError:
                        pass
            
            result = {
                'total_size': total_size,
                'total_size_mb': total_size / (1024 * 1024),
                'total_size_gb': total_size / (1024 * 1024 * 1024),
                'file_count': file_count,
                'folder_count': folder_count,
                'data_dir': self.data_dir
            }
            
            self.analysis_complete.emit(result)
        except Exception as e:
            logger.error(f"Storage analysis error: {e}")
            self.analysis_complete.emit({})


class DataManagementTab(QWidget):
    """
    Centralized Data Management Tab
    
    Features:
    - Export data (CSV, JSON, Excel)
    - Import data
    - Backup management
    - Data cleanup and optimization
    - Storage usage monitoring
    """
    
    # Signal for data export/import events
    data_exported = Signal(str, str)  # export_type, file_path
    data_imported = Signal(str)  # import_type
    
    def __init__(self, parent=None, historical_data_manager=None, vehicle_manager=None):
        super().__init__(parent)
        self.data_dir = str(CONFIG.DATA_DIR) if CONFIG else "./data"
        self.storage_analyzer = None
        self.historical_data_manager = historical_data_manager
        self.vehicle_manager = vehicle_manager

        # Initialize backup manager
        self.backup_manager = None
        if BACKUP_MANAGER_AVAILABLE:
            try:
                self.backup_manager = BackupManager()
                logger.info("Backup manager initialized")
            except Exception as e:
                logger.error(f"Failed to initialize backup manager: {e}")

        # Initialize data export system (will be initialized when needed with managers)
        self.data_export = None

        self._setup_ui()
        self._start_storage_monitoring()
        self._load_backup_history()
    
    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(15)
        
        # Header
        header = QHBoxLayout()
        title = QLabel("Data Management")
        title.setStyleSheet("font-size: 18px; font-weight: 700; color: #F0F6FC;")
        
        self.refresh_btn = QPushButton("⟳ Refresh")
        self.refresh_btn.setStyleSheet(self._get_button_style('secondary'))
        self.refresh_btn.clicked.connect(self._refresh_storage_info)
        
        header.addWidget(title)
        header.addStretch()
        header.addWidget(self.refresh_btn)
        layout.addLayout(header)
        
        # Main tab widget
        self.tabs = QTabWidget()
        self.tabs.setStyleSheet("""
            QTabWidget::pane {
                border: 1px solid #30363D;
                border-radius: 8px;
                background-color: #161B22;
                padding: 5px;
            }
            QTabBar::tab {
                background-color: #21262D;
                color: #8B949E;
                padding: 10px 18px;
                margin-right: 2px;
                border-top-left-radius: 6px;
                border-top-right-radius: 6px;
                border: 1px solid #30363D;
                border-bottom: none;
                font-weight: 500;
            }
            QTabBar::tab:selected {
                background-color: #C40000;
                color: #F0F6FC;
                font-weight: 600;
            }
            QTabBar::tab:hover:!selected {
                background-color: #30363D;
                color: #F0F6FC;
            }
        """)
        
        # Tab 1: Storage Overview
        storage_tab = self._create_storage_tab()
        self.tabs.addTab(storage_tab, "💾 Storage")
        
        # Tab 2: Export Data
        export_tab = self._create_export_tab()
        self.tabs.addTab(export_tab, "📤 Export")
        
        # Tab 3: Import Data
        import_tab = self._create_import_tab()
        self.tabs.addTab(import_tab, "📥 Import")
        
        # Tab 4: Backup Management
        backup_tab = self._create_backup_tab()
        self.tabs.addTab(backup_tab, "🗄️ Backups")
        
        # Tab 5: Data Cleanup
        cleanup_tab = self._create_cleanup_tab()
        self.tabs.addTab(cleanup_tab, "🧹 Cleanup")
        
        layout.addWidget(self.tabs, 1)
    
    def _create_storage_tab(self) -> QWidget:
        """Create storage overview tab"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(15)
        
        # Storage summary cards
        summary_layout = QHBoxLayout()
        
        # Total size card
        size_card = QGroupBox("Total Storage")
        size_layout = QFormLayout(size_card)
        self.total_size_label = QLabel("0 MB")
        self.total_size_label.setStyleSheet("font-size: 24px; font-weight: 700; color: #C40000;")
        size_layout.addRow("Used:", self.total_size_label)
        summary_layout.addWidget(size_card)
        
        # File count card
        files_card = QGroupBox("Files")
        files_layout = QFormLayout(files_card)
        self.file_count_label = QLabel("0")
        self.file_count_label.setStyleSheet("font-size: 24px; font-weight: 700; color: #4CAF50;")
        files_layout.addRow("Total:", self.file_count_label)
        summary_layout.addWidget(files_card)
        
        # Folder count card
        folders_card = QGroupBox("Folders")
        folders_layout = QFormLayout(folders_card)
        self.folder_count_label = QLabel("0")
        self.folder_count_label.setStyleSheet("font-size: 24px; font-weight: 700; color: #0DCAF0;")
        folders_layout.addRow("Total:", self.folder_count_label)
        summary_layout.addWidget(folders_card)
        
        layout.addLayout(summary_layout)
        
        # Storage breakdown by category
        breakdown_group = QGroupBox("Storage Breakdown")
        breakdown_layout = QVBoxLayout(breakdown_group)
        
        self.storage_table = QTableWidget()
        self.storage_table.setColumnCount(4)
        self.storage_table.setHorizontalHeaderLabels(["Category", "Files", "Size", "%"])
        self.storage_table.horizontalHeader().setStretchLastSection(True)
        self.storage_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.storage_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self._apply_table_styling(self.storage_table)
        
        breakdown_layout.addWidget(self.storage_table)
        layout.addWidget(breakdown_group)
        
        # Progress bar
        progress_group = QGroupBox("Storage Usage")
        progress_layout = QVBoxLayout(progress_group)
        
        self.storage_progress = QProgressBar()
        self.storage_progress.setMaximum(100)
        self.storage_progress.setValue(0)
        self.storage_progress.setTextVisible(True)
        self.storage_progress.setFormat("%p% used")
        self.storage_progress.setStyleSheet("""
            QProgressBar {
                border:1px solid #30363D;
                border-radius: 3px;
                background-color: #161B22;
                text-align: center;
                height: 25px;
                font-size: 12px;
                font-weight: 600;
            }
            QProgressBar::chunk {
                background-color: #C40000;
                border-radius: 3px;
            }
        """)
        
        progress_layout.addWidget(self.storage_progress)
        layout.addWidget(progress_group)
        
        layout.addStretch()
        
        return widget
    
    def _create_export_tab(self) -> QWidget:
        """Create data export tab"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(15)
        
        # Export options
        options_group = QGroupBox("Export Options")
        options_layout = QFormLayout(options_group)
        
        self.export_type_combo = QComboBox()
        self.export_type_combo.addItems([
            "CSV (Comma Separated)",
            "JSON (JavaScript Object Notation)",
            "Excel (.xlsx)",
            "All Data (ZIP Archive)"
        ])
        options_layout.addRow("Format:", self.export_type_combo)
        
        self.export_profile_combo = QComboBox()
        self.export_profile_combo.addItem("All Profiles")
        options_layout.addRow("Profile:", self.export_profile_combo)
        
        self.date_range_combo = QComboBox()
        self.date_range_combo.addItems([
            "All Time",
            "Last 7 Days",
            "Last 30 Days",
            "Last 90 Days",
            "Custom Range"
        ])
        options_layout.addRow("Date Range:", self.date_range_combo)
        
        self.include_obd_check = QCheckBox("Include OBD Data")
        self.include_obd_check.setChecked(True)
        options_layout.addRow("", self.include_obd_check)
        
        self.include_service_check = QCheckBox("Include Service History")
        self.include_service_check.setChecked(True)
        options_layout.addRow("", self.include_service_check)
        
        self.include_predictions_check = QCheckBox("Include AI Predictions")
        self.include_predictions_check.setChecked(True)
        options_layout.addRow("", self.include_predictions_check)
        
        layout.addWidget(options_group)
        
        # Export button
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        self.export_btn = QPushButton("📤 Export Data")
        self.export_btn.setStyleSheet(self._get_button_style('primary'))
        self.export_btn.clicked.connect(self._export_data)
        button_layout.addWidget(self.export_btn)
        
        layout.addLayout(button_layout)
        layout.addStretch()
        
        return widget
    
    def _create_import_tab(self) -> QWidget:
        """Create data import tab"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(15)
        
        # Import options
        options_group = QGroupBox("Import Options")
        options_layout = QFormLayout(options_group)
        
        self.import_file_edit = QLineEdit()
        self.import_file_edit.setPlaceholderText("Select file to import...")
        self.import_file_edit.setReadOnly(True)
        options_layout.addRow("File:", self.import_file_edit)
        
        browse_btn = QPushButton("Browse...")
        browse_btn.setStyleSheet(self._get_button_style('secondary'))
        browse_btn.clicked.connect(self._browse_import_file)
        options_layout.addRow("", browse_btn)
        
        self.import_type_combo = QComboBox()
        self.import_type_combo.addItems([
            "Auto-detect",
            "CSV (Comma Separated)",
            "JSON (JavaScript Object Notation)",
            "Excel (.xlsx)"
        ])
        options_layout.addRow("Format:", self.import_type_combo)
        
        self.import_profile_combo = QComboBox()
        self.import_profile_combo.addItem("Create New Profile")
        options_layout.addRow("Target Profile:", self.import_profile_combo)
        
        self.overwrite_check = QCheckBox("Overwrite Existing Data")
        self.overwrite_check.setChecked(False)
        options_layout.addRow("", self.overwrite_check)
        
        layout.addWidget(options_group)
        
        # Import button
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        self.import_btn = QPushButton("📥 Import Data")
        self.import_btn.setStyleSheet(self._get_button_style('success'))
        self.import_btn.clicked.connect(self._import_data)
        button_layout.addWidget(self.import_btn)
        
        layout.addLayout(button_layout)
        layout.addStretch()
        
        return widget
    
    def _create_backup_tab(self) -> QWidget:
        """Create backup management tab"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(15)
        
        # Backup controls
        controls_group = QGroupBox("Backup Controls")
        controls_layout = QFormLayout(controls_group)
        
        self.backup_profile_combo = QComboBox()
        self.backup_profile_combo.addItem("All Profiles")
        controls_layout.addRow("Profile:", self.backup_profile_combo)
        
        self.backup_type_combo = QComboBox()
        self.backup_type_combo.addItems([
            "Full Backup (All Data)",
            "OBD Data Only",
            "Service History Only",
            "AI Models Only",
            "Settings Only"
        ])
        controls_layout.addRow("Type:", self.backup_type_combo)
        
        self.auto_backup_check = QCheckBox("Enable Automatic Daily Backups")
        self.auto_backup_check.setChecked(True)
        controls_layout.addRow("", self.auto_backup_check)
        
        self.backup_time_edit = QLineEdit("02:00")
        self.backup_time_edit.setPlaceholderText("HH:MM")
        controls_layout.addRow("Backup Time:", self.backup_time_edit)
        
        layout.addWidget(controls_group)
        
        # Backup buttons
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        self.create_backup_btn = QPushButton("🗄️ Create Backup")
        self.create_backup_btn.setStyleSheet(self._get_button_style('success'))
        self.create_backup_btn.clicked.connect(self._create_backup)
        button_layout.addWidget(self.create_backup_btn)
        
        self.restore_backup_btn = QPushButton("📂 Restore Backup")
        self.restore_backup_btn.setStyleSheet(self._get_button_style('info'))
        self.restore_backup_btn.clicked.connect(self._restore_backup)
        button_layout.addWidget(self.restore_backup_btn)
        
        layout.addLayout(button_layout)
        
        # Backup history
        history_group = QGroupBox("Backup History")
        history_layout = QVBoxLayout(history_group)
        
        self.backup_table = QTableWidget()
        self.backup_table.setColumnCount(5)
        self.backup_table.setHorizontalHeaderLabels(["Date", "Type", "Profile", "Size", "Actions"])
        self.backup_table.horizontalHeader().setStretchLastSection(True)
        self.backup_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.backup_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self._apply_table_styling(self.backup_table)
        
        history_layout.addWidget(self.backup_table)
        layout.addWidget(history_group)
        
        return widget
    
    def _create_cleanup_tab(self) -> QWidget:
        """Create data cleanup tab"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(15)
        
        # Cleanup options
        options_group = QGroupBox("Cleanup Options")
        options_layout = QVBoxLayout(options_group)
        
        self.cleanup_obd_check = QCheckBox("Clean Old OBD Data (older than 90 days)")
        self.cleanup_obd_check.setChecked(False)
        options_layout.addWidget(self.cleanup_obd_check)
        
        self.cleanup_temp_check = QCheckBox("Clean Temporary Files")
        self.cleanup_temp_check.setChecked(True)
        options_layout.addWidget(self.cleanup_temp_check)
        
        self.cleanup_logs_check = QCheckBox("Clean Old Log Files (older than 30 days)")
        self.cleanup_logs_check.setChecked(False)
        options_layout.addWidget(self.cleanup_logs_check)
        
        self.cleanup_cache_check = QCheckBox("Clear AI Model Cache")
        self.cleanup_cache_check.setChecked(False)
        options_layout.addWidget(self.cleanup_cache_check)
        
        self.cleanup_orphaned_check = QCheckBox("Remove Orphaned Files")
        self.cleanup_orphaned_check.setChecked(False)
        options_layout.addWidget(self.cleanup_orphaned_check)
        
        layout.addWidget(options_group)
        
        # Cleanup button
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        self.cleanup_btn = QPushButton("🧹 Run Cleanup")
        self.cleanup_btn.setStyleSheet(self._get_button_style('warning'))
        self.cleanup_btn.clicked.connect(self._run_cleanup)
        button_layout.addWidget(self.cleanup_btn)
        
        layout.addLayout(button_layout)
        
        # Cleanup progress
        self.cleanup_progress = QProgressBar()
        self.cleanup_progress.setVisible(False)
        layout.addWidget(self.cleanup_progress)
        
        # Cleanup results
        self.cleanup_results = QLabel("")
        self.cleanup_results.setWordWrap(True)
        self.cleanup_results.setStyleSheet("color: #8B949E; font-size: 12px;")
        layout.addWidget(self.cleanup_results)
        
        layout.addStretch()
        
        return widget
    
    def _get_button_style(self, style_type: str = 'primary') -> str:
        """Get consistent button stylesheet based on style type."""
        styles = {
            'primary': """
                QPushButton {
                    background-color: #C40000;
                    color: #FFFFFF;
                    border: none;
                    padding: 12px 24px;
                    font-size: 14px;
                    font-weight: 600;
                    border-radius: 6px;
                }
                QPushButton:hover {
                    background-color: #E53935;
                }
                QPushButton:pressed {
                    background-color: #A00000;
                }
                QPushButton:disabled {
                    background-color: #30363D;
                    color: #8B949E;
                }
            """,
            'secondary': """
                QPushButton {
                    background-color: #21262D;
                    color: #F0F6FC;
                    border: 1px solid #30363D;
                    padding: 10px 20px;
                    font-size: 14px;
                    font-weight: 600;
                    border-radius: 6px;
                }
                QPushButton:hover {
                    background-color: #30363D;
                }
                QPushButton:disabled {
                    background-color: #161B22;
                    color: #8B949E;
                }
            """,
            'danger': """
                QPushButton {
                    background-color: #C40000;
                    color: #FFFFFF;
                    border: none;
                    padding: 10px 20px;
                    font-size: 14px;
                    font-weight: 600;
                    border-radius: 6px;
                }
                QPushButton:hover {
                    background-color: #E53935;
                }
                QPushButton:pressed {
                    background-color: #A00000;
                }
                QPushButton:disabled {
                    background-color: #30363D;
                    color: #8B949E;
                }
            """,
            'success': """
                QPushButton {
                    background-color: #4CAF50;
                    color: #FFFFFF;
                    border: none;
                    padding: 10px 20px;
                    font-size: 14px;
                    font-weight: 600;
                    border-radius: 6px;
                }
                QPushButton:hover {
                    background-color: #45a049;
                }
                QPushButton:pressed {
                    background-color: #388E3C;
                }
                QPushButton:disabled {
                    background-color: #30363D;
                    color: #8B949E;
                }
            """,
            'warning': """
                QPushButton {
                    background-color: #FFC107;
                    color: #FFFFFF;
                    border: none;
                    padding: 10px 20px;
                    font-size: 14px;
                    font-weight: 600;
                    border-radius: 6px;
                }
                QPushButton:hover {
                    background-color: #FFD54F;
                }
                QPushButton:pressed {
                    background-color: #FFA000;
                }
                QPushButton:disabled {
                    background-color: #30363D;
                    color: #8B949E;
                }
            """,
            'info': """
                QPushButton {
                    background-color: #2196F3;
                    color: #FFFFFF;
                    border: none;
                    padding: 10px 20px;
                    font-size: 14px;
                    font-weight: 600;
                    border-radius: 6px;
                }
                QPushButton:hover {
                    background-color: #1976D2;
                }
                QPushButton:pressed {
                    background-color: #0D47A1;
                }
                QPushButton:disabled {
                    background-color: #30363D;
                    color: #8B949E;
                }
            """
        }
        return styles.get(style_type, styles['primary'])

    def _apply_table_styling(self, table_widget):
        """Apply consistent table styling to a QTableWidget"""
        table_widget.setStyleSheet("""
            QTableWidget {
                background-color: #21262D;
                color: #F0F6FC;
                border: 1px solid #30363D;
                border-radius: 8px;
                gridline-color: #30363D;
            }
            QTableWidget::item {
                padding: 10px;
                border-bottom: 1px solid #30363D;
            }
            QTableWidget::item:selected {
                background-color: #C40000;
                color: #FFFFFF;
            }
            QTableWidget::item:hover {
                background-color: #30363D;
            }
            QHeaderView::section {
                background-color: #161B22;
                color: #F0F6FC;
                padding: 12px;
                border: none;
                border-bottom: 2px solid #C40000;
                font-weight: 600;
            }
            QScrollBar:vertical {
                background-color: #161B22;
                width: 12px;
                border-radius: 6px;
            }
            QScrollBar::handle:vertical {
                background-color: #484F58;
                border-radius: 6px;
                min-height: 30px;
            }
            QScrollBar::handle:vertical:hover {
                background-color: #6E7681;
            }
        """)
    
    def _start_storage_monitoring(self):
        """Start periodic storage monitoring"""
        self.storage_timer = QTimer(self)
        self.storage_timer.timeout.connect(self._refresh_storage_info)
        self.storage_timer.start(60000)  # Update every minute
        
        # Initial refresh
        self._refresh_storage_info()
    
    def _refresh_storage_info(self):
        """Refresh storage information"""
        try:
            # Start storage analyzer in background
            self.storage_analyzer = StorageAnalyzer(self.data_dir)
            self.storage_analyzer.analysis_complete.connect(self._update_storage_display)
            self.storage_analyzer.start()
        except Exception as e:
            logger.error(f"Error refreshing storage info: {e}")
    
    def _update_storage_display(self, result: Dict[str, Any]):
        """Update storage display with analysis results"""
        try:
            total_size = result.get('total_size_mb', 0)
            file_count = result.get('file_count', 0)
            folder_count = result.get('folder_count', 0)
            
            # Update labels
            if total_size >= 1024:
                size_text = f"{total_size / 1024:.2f} GB"
            else:
                size_text = f"{total_size:.2f} MB"
            
            self.total_size_label.setText(size_text)
            self.file_count_label.setText(str(file_count))
            self.folder_count_label.setText(str(folder_count))
            
            # Update progress bar (assuming 10 GB limit)
            max_gb = 10.0
            used_gb = result.get('total_size_gb', 0)
            percentage = min(100, int((used_gb / max_gb) * 100))
            
            self.storage_progress.setValue(percentage)
            
            # Update storage breakdown table
            self._update_storage_breakdown()
            
        except Exception as e:
            logger.error(f"Error updating storage display: {e}")
    
    def _update_storage_breakdown(self):
        """Update storage breakdown table"""
        try:
            self.storage_table.setRowCount(0)
            
            # Analyze each subdirectory
            categories = [
                ("OBD Data", "obd_data"),
                ("Service History", "service_history"),
                ("AI Models", "ai_models"),
                ("Reports", "reports"),
                ("Backups", "backups"),
                ("Logs", "logs"),
                ("Other", "other")
            ]
            
            total_size = 0
            
            for category_name, dir_name in categories:
                dir_path = os.path.join(self.data_dir, dir_name)
                
                if os.path.exists(dir_path):
                    size = self._get_directory_size(dir_path)
                    files = self._count_files(dir_path)
                    total_size += size
                else:
                    size = 0
                    files = 0
                
                row = self.storage_table.rowCount()
                self.storage_table.insertRow(row)
                self.storage_table.setItem(row, 0, QTableWidgetItem(category_name))
                self.storage_table.setItem(row, 1, QTableWidgetItem(str(files)))
                self.storage_table.setItem(row, 2, QTableWidgetItem(f"{size / (1024*1024):.2f} MB"))
                
                if total_size > 0:
                    percentage = (size / total_size) * 100
                    self.storage_table.setItem(row, 3, QTableWidgetItem(f"{percentage:.1f}%"))
                else:
                    self.storage_table.setItem(row, 3, QTableWidgetItem("0%"))
            
            self.storage_table.resizeColumnsToContents()
            
        except Exception as e:
            logger.error(f"Error updating storage breakdown: {e}")
    
    def _get_directory_size(self, path: str) -> int:
        """Get total size of directory in bytes"""
        total_size = 0
        try:
            for root, dirs, files in os.walk(path):
                for file in files:
                    file_path = os.path.join(root, file)
                    try:
                        total_size += os.path.getsize(file_path)
                    except OSError:
                        pass
        except Exception:
            pass
        return total_size
    
    def _count_files(self, path: str) -> int:
        """Count files in directory"""
        count = 0
        try:
            for root, dirs, files in os.walk(path):
                count += len(files)
        except Exception:
            pass
        return count
    
    def _export_data(self):
        """Export data based on selected options"""
        try:
            export_type = self.export_type_combo.currentText()
            profile = self.export_profile_combo.currentText()
            date_range = self.date_range_combo.currentText()

            # Determine what to export
            include_obd = self.include_obd_check.isChecked()
            include_service = self.include_service_check.isChecked()
            include_predictions = self.include_predictions_check.isChecked()

            # Determine format
            if "CSV" in export_type:
                export_format = 'csv'
            elif "JSON" in export_type:
                export_format = 'json'
            elif "Excel" in export_type:
                export_format = 'excel'
            else:
                export_format = 'zip'

            # Calculate date range
            end_date = datetime.now()
            if "7 Days" in date_range:
                start_date = end_date - timedelta(days=7)
            elif "30 Days" in date_range:
                start_date = end_date - timedelta(days=30)
            elif "90 Days" in date_range:
                start_date = end_date - timedelta(days=90)
            else:
                start_date = None  # All time

            # Get export directory
            export_dir = str(CONFIG.DATA_DIR / "exports") if CONFIG else "./data/exports"
            os.makedirs(export_dir, exist_ok=True)

            # Initialize data export system if not done
            if DATA_EXPORT_AVAILABLE and self.data_export is None and self.historical_data_manager:
                self.data_export = DataExportSystem(
                    self.historical_data_manager,
                    self.vehicle_manager
                )

            exported_files = []
            errors = []

            # Export based on type
            if export_format == 'zip' or profile == "All Profiles":
                # Export all data as ZIP archive
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                zip_filename = f"predict_full_export_{timestamp}.zip"
                zip_path = os.path.join(export_dir, zip_filename)

                with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    # Export OBD data
                    if include_obd and self.historical_data_manager:
                        try:
                            obd_data = self._gather_obd_data(start_date, end_date)
                            if obd_data:
                                obd_json = json.dumps(obd_data, indent=2, default=str)
                                zipf.writestr('obd_data.json', obd_json)
                        except Exception as e:
                            errors.append(f"OBD data: {e}")

                    # Export service history
                    if include_service:
                        try:
                            service_data = self._gather_service_history()
                            if service_data:
                                service_json = json.dumps(service_data, indent=2, default=str)
                                zipf.writestr('service_history.json', service_json)
                        except Exception as e:
                            errors.append(f"Service history: {e}")

                    # Export AI predictions
                    if include_predictions:
                        try:
                            predictions_data = self._gather_predictions()
                            if predictions_data:
                                predictions_json = json.dumps(predictions_data, indent=2, default=str)
                                zipf.writestr('ai_predictions.json', predictions_json)
                        except Exception as e:
                            errors.append(f"Predictions: {e}")

                exported_files.append(zip_path)

            else:
                # Export specific format
                if self.data_export and self.historical_data_manager:
                    # Get profile ID (for now, use profile 1 as default)
                    profile_id = 1
                    profile_name = profile if profile != "All Profiles" else "default"

                    result = self.data_export.export_obd_data(
                        profile_id=profile_id,
                        profile_name=profile_name,
                        start_date=start_date,
                        end_date=end_date,
                        format=export_format
                    )

                    if result.get('success'):
                        exported_files.append(result['file_path'])
                    else:
                        errors.append(result.get('error', 'Unknown error'))
                else:
                    # Fallback: Simple JSON/CSV export
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

                    if export_format == 'json':
                        filename = f"predict_export_{timestamp}.json"
                        file_path = os.path.join(export_dir, filename)
                        data = self._gather_all_data(include_obd, include_service, include_predictions)
                        with open(file_path, 'w', encoding='utf-8') as f:
                            json.dump(data, f, indent=2, default=str)
                        exported_files.append(file_path)

                    elif export_format == 'csv':
                        filename = f"predict_export_{timestamp}.csv"
                        file_path = os.path.join(export_dir, filename)
                        data = self._gather_obd_data(start_date, end_date)
                        if data:
                            self._write_csv(file_path, data)
                        exported_files.append(file_path)

            # Show result
            if exported_files:
                file_list = "\n".join(exported_files)
                message = f"✅ Export completed successfully!\n\nFiles created:\n{file_list}"
                if errors:
                    message += f"\n\n⚠️ Some errors occurred:\n" + "\n".join(errors)

                QMessageBox.information(self, "Export Complete", message)

                # Emit signal
                self.data_exported.emit(export_type, exported_files[0])
                logger.info(f"Data export completed: {exported_files}")
            else:
                QMessageBox.warning(
                    self,
                    "Export Warning",
                    f"No data was exported.\n\nErrors:\n" + "\n".join(errors) if errors else "No data available to export."
                )

        except Exception as e:
            logger.error(f"Error exporting data: {e}")
            QMessageBox.critical(self, "Export Error", f"Failed to export data:\n{e}")
    
    def _import_data(self):
        """Import data from selected file"""
        try:
            file_path = self.import_file_edit.text()

            if not file_path or not os.path.exists(file_path):
                QMessageBox.warning(self, "No File Selected", "Please select a file to import.")
                return

            import_type = self.import_type_combo.currentText()
            target_profile = self.import_profile_combo.currentText()
            overwrite = self.overwrite_check.isChecked()

            # Confirm overwrite
            if overwrite:
                confirm = QMessageBox.question(
                    self,
                    "Confirm Overwrite",
                    "This will overwrite existing data. Are you sure you want to continue?",
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.No
                )
                if confirm != QMessageBox.Yes:
                    return

            # Detect file type
            file_ext = os.path.splitext(file_path)[1].lower()

            if import_type == "Auto-detect":
                if file_ext == '.csv':
                    import_format = 'csv'
                elif file_ext == '.json':
                    import_format = 'json'
                elif file_ext in ['.xlsx', '.xls']:
                    import_format = 'excel'
                elif file_ext == '.zip':
                    import_format = 'zip'
                else:
                    QMessageBox.warning(self, "Unknown Format", f"Cannot auto-detect format for: {file_ext}")
                    return
            else:
                if "CSV" in import_type:
                    import_format = 'csv'
                elif "JSON" in import_type:
                    import_format = 'json'
                else:
                    import_format = 'excel'

            imported_records = 0
            errors = []

            # Import based on format
            if import_format == 'zip':
                # Extract and import ZIP archive
                import_dir = str(CONFIG.DATA_DIR / "imports") if CONFIG else "./data/imports"
                os.makedirs(import_dir, exist_ok=True)

                with zipfile.ZipFile(file_path, 'r') as zipf:
                    zipf.extractall(import_dir)
                    for name in zipf.namelist():
                        if name.endswith('.json'):
                            extracted_path = os.path.join(import_dir, name)
                            try:
                                with open(extracted_path, 'r', encoding='utf-8') as f:
                                    data = json.load(f)
                                    if isinstance(data, list):
                                        imported_records += len(data)
                                    elif isinstance(data, dict):
                                        imported_records += 1
                            except Exception as e:
                                errors.append(f"{name}: {e}")

            elif import_format == 'json':
                with open(file_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    if isinstance(data, list):
                        imported_records = len(data)
                    elif isinstance(data, dict):
                        imported_records = 1
                    # Store imported data in historical data manager if available
                    if self.historical_data_manager and isinstance(data, list):
                        for record in data:
                            try:
                                self.historical_data_manager.store_reading(record)
                            except Exception as e:
                                errors.append(str(e))

            elif import_format == 'csv':
                with open(file_path, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        imported_records += 1
                        if self.historical_data_manager:
                            try:
                                self.historical_data_manager.store_reading(row)
                            except Exception as e:
                                errors.append(str(e))

            elif import_format == 'excel':
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(file_path)
                    ws = wb.active
                    headers = [cell.value for cell in ws[1]]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        record = dict(zip(headers, row))
                        imported_records += 1
                        if self.historical_data_manager:
                            try:
                                self.historical_data_manager.store_reading(record)
                            except Exception as e:
                                errors.append(str(e))
                except ImportError:
                    QMessageBox.warning(self, "Excel Not Available", "openpyxl is required for Excel import.")
                    return

            # Show result
            if imported_records > 0:
                message = f"✅ Import completed successfully!\n\nRecords imported: {imported_records}"
                if errors:
                    message += f"\n\n⚠️ Errors ({len(errors)}):\n" + "\n".join(errors[:5])
                    if len(errors) > 5:
                        message += f"\n... and {len(errors) - 5} more errors"
                QMessageBox.information(self, "Import Complete", message)
                self.data_imported.emit(import_format)
                logger.info(f"Data import completed: {imported_records} records from {file_path}")
            else:
                QMessageBox.warning(self, "Import Warning", "No records were imported from the file.")

        except Exception as e:
            logger.error(f"Error importing data: {e}")
            QMessageBox.critical(self, "Import Error", f"Failed to import data:\n{e}")
    
    def _browse_import_file(self):
        """Browse for import file"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Import File",
            "",
            "Data Files (*.csv *.json *.xlsx *.zip);;All Files (*.*)"
        )
        
        if file_path:
            self.import_file_edit.setText(file_path)
    
    def _create_backup(self):
        """Create backup of selected data"""
        try:
            profile = self.backup_profile_combo.currentText()
            backup_type_text = self.backup_type_combo.currentText()

            # Disable button during backup
            self.create_backup_btn.setEnabled(False)
            self.create_backup_btn.setText("Creating backup...")

            # Determine backup type
            if "Full" in backup_type_text:
                backup_type = 'daily'
                description = f"Manual full backup - {profile}"
            elif "OBD" in backup_type_text:
                backup_type = 'daily'
                description = f"OBD data backup - {profile}"
            elif "Service" in backup_type_text:
                backup_type = 'daily'
                description = f"Service history backup - {profile}"
            elif "AI" in backup_type_text:
                backup_type = 'daily'
                description = f"AI models backup - {profile}"
            else:
                backup_type = 'daily'
                description = f"Settings backup - {profile}"

            # Create backup using backup manager
            if self.backup_manager:
                backup_path = self.backup_manager.create_backup(
                    backup_type=backup_type,
                    description=description
                )

                if backup_path:
                    # Get file size
                    size_mb = os.path.getsize(backup_path) / (1024 * 1024)

                    QMessageBox.information(
                        self,
                        "Backup Complete",
                        f"✅ Backup created successfully!\n\n"
                        f"File: {os.path.basename(backup_path)}\n"
                        f"Size: {size_mb:.2f} MB\n"
                        f"Location: {os.path.dirname(backup_path)}"
                    )

                    # Refresh backup history
                    self._load_backup_history()

                    logger.info(f"Backup created: {backup_path}")
                else:
                    QMessageBox.warning(
                        self,
                        "Backup Warning",
                        "Backup was not created. Check logs for details."
                    )
            else:
                # Fallback: Simple ZIP backup
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                backup_dir = str(CONFIG.BACKUPS_DIR) if CONFIG else "./backups"
                os.makedirs(backup_dir, exist_ok=True)

                backup_filename = f"backup_{timestamp}.zip"
                backup_path = os.path.join(backup_dir, backup_filename)

                with zipfile.ZipFile(backup_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    # Backup data directory
                    data_dir = self.data_dir
                    for root, dirs, files in os.walk(data_dir):
                        for file in files:
                            file_path = os.path.join(root, file)
                            arcname = os.path.relpath(file_path, data_dir)
                            zipf.write(file_path, arcname)

                size_mb = os.path.getsize(backup_path) / (1024 * 1024)
                QMessageBox.information(
                    self,
                    "Backup Complete",
                    f"✅ Backup created!\n\nFile: {backup_filename}\nSize: {size_mb:.2f} MB"
                )

                self._load_backup_history()

        except Exception as e:
            logger.error(f"Error creating backup: {e}")
            QMessageBox.critical(self, "Backup Error", f"Failed to create backup:\n{e}")
        finally:
            # Re-enable button
            self.create_backup_btn.setEnabled(True)
            self.create_backup_btn.setText("🗄️ Create Backup")
    
    def _restore_backup(self):
        """Restore data from selected backup"""
        try:
            # Get selected backup from table
            selected_rows = self.backup_table.selectedItems()
            if not selected_rows:
                # Ask user to browse for backup file
                backup_path, _ = QFileDialog.getOpenFileName(
                    self,
                    "Select Backup File",
                    str(CONFIG.BACKUPS_DIR) if CONFIG else "./backups",
                    "Backup Files (*.zip);;All Files (*.*)"
                )
                if not backup_path:
                    return
            else:
                # Get backup path from selected row
                row = self.backup_table.currentRow()
                if row < 0:
                    QMessageBox.warning(self, "No Selection", "Please select a backup to restore.")
                    return
                # Reconstruct backup path from table data
                date_item = self.backup_table.item(row, 0)
                backup_type_item = self.backup_table.item(row, 1)
                if date_item and backup_type_item:
                    backup_type = backup_type_item.text().lower()
                    backup_dir = os.path.join(
                        str(CONFIG.BACKUPS_DIR) if CONFIG else "./backups",
                        backup_type
                    )
                    # Find matching backup file
                    backup_path = None
                    if os.path.exists(backup_dir):
                        for f in os.listdir(backup_dir):
                            if f.endswith('.zip'):
                                backup_path = os.path.join(backup_dir, f)
                                break
                    if not backup_path:
                        QMessageBox.warning(self, "Backup Not Found", "Could not locate the backup file.")
                        return
                else:
                    QMessageBox.warning(self, "Selection Error", "Invalid backup selection.")
                    return

            # Confirm restore
            confirm = QMessageBox.warning(
                self,
                "Confirm Restore",
                f"⚠️ WARNING: This will overwrite current data!\n\n"
                f"Backup file: {os.path.basename(backup_path)}\n\n"
                f"Are you sure you want to restore this backup?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )

            if confirm != QMessageBox.Yes:
                return

            # Create a backup of current data before restore
            pre_restore_backup = None
            if self.backup_manager:
                pre_restore_backup = self.backup_manager.create_backup(
                    backup_type='daily',
                    description='Pre-restore safety backup'
                )

            # Perform restore
            if self.backup_manager:
                success = self.backup_manager.restore_backup(backup_path)
            else:
                # Fallback: Manual restore
                restore_dir = self.data_dir
                with zipfile.ZipFile(backup_path, 'r') as zipf:
                    zipf.extractall(restore_dir)
                success = True

            if success:
                message = "✅ Backup restored successfully!\n\nPlease restart the application for changes to take effect."
                if pre_restore_backup:
                    message += f"\n\nA safety backup was created at:\n{pre_restore_backup}"
                QMessageBox.information(self, "Restore Complete", message)
                logger.info(f"Backup restored: {backup_path}")
            else:
                QMessageBox.critical(self, "Restore Failed", "Failed to restore backup. Check logs for details.")

        except Exception as e:
            logger.error(f"Error restoring backup: {e}")
            QMessageBox.critical(self, "Restore Error", f"Failed to restore backup:\n{e}")
    
    def _run_cleanup(self):
        """Run data cleanup operations"""
        try:
            cleanup_items = []

            if self.cleanup_obd_check.isChecked():
                cleanup_items.append(("Old OBD Data", self._cleanup_old_obd_data))

            if self.cleanup_temp_check.isChecked():
                cleanup_items.append(("Temporary Files", self._cleanup_temp_files))

            if self.cleanup_logs_check.isChecked():
                cleanup_items.append(("Old Log Files", self._cleanup_old_logs))

            if self.cleanup_cache_check.isChecked():
                cleanup_items.append(("AI Model Cache", self._cleanup_ai_cache))

            if self.cleanup_orphaned_check.isChecked():
                cleanup_items.append(("Orphaned Files", self._cleanup_orphaned_files))

            if not cleanup_items:
                QMessageBox.information(self, "No Selection", "Please select at least one cleanup option.")
                return

            # Confirm cleanup
            confirm = QMessageBox.question(
                self,
                "Confirm Cleanup",
                f"This will clean up the following:\n\n" +
                "\n".join(f"  • {item[0]}" for item in cleanup_items) +
                "\n\nThis action cannot be undone. Continue?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )

            if confirm != QMessageBox.Yes:
                return

            # Show progress
            self.cleanup_progress.setVisible(True)
            self.cleanup_progress.setRange(0, len(cleanup_items))
            self.cleanup_results.setText("Running cleanup operations...")
            self.cleanup_btn.setEnabled(False)

            total_files_deleted = 0
            total_space_freed = 0
            results = []

            # Run cleanup operations
            for i, (name, cleanup_func) in enumerate(cleanup_items):
                self.cleanup_progress.setValue(i)
                self.cleanup_results.setText(f"Cleaning: {name}...")

                try:
                    files_deleted, space_freed = cleanup_func()
                    total_files_deleted += files_deleted
                    total_space_freed += space_freed
                    results.append(f"  ✅ {name}: {files_deleted} files ({space_freed / (1024*1024):.2f} MB)")
                except Exception as e:
                    results.append(f"  ❌ {name}: Error - {e}")
                    logger.error(f"Cleanup error for {name}: {e}")

            self.cleanup_progress.setValue(len(cleanup_items))
            self.cleanup_progress.setVisible(False)

            # Show results
            space_freed_mb = total_space_freed / (1024 * 1024)
            result_text = (
                f"✅ Cleanup complete!\n\n"
                f"Files deleted: {total_files_deleted}\n"
                f"Space freed: {space_freed_mb:.2f} MB\n\n"
                f"Details:\n" + "\n".join(results)
            )
            self.cleanup_results.setText(result_text)

            logger.info(f"Data cleanup completed: {total_files_deleted} files, {space_freed_mb:.2f} MB freed")

            # Refresh storage info
            self._refresh_storage_info()

        except Exception as e:
            logger.error(f"Error running cleanup: {e}")
            self.cleanup_progress.setVisible(False)
            QMessageBox.critical(self, "Cleanup Error", f"Failed to run cleanup:\n{e}")
        finally:
            self.cleanup_btn.setEnabled(True)

    def _cleanup_old_obd_data(self) -> tuple:
        """Clean up OBD data older than 90 days"""
        files_deleted = 0
        space_freed = 0
        cutoff_date = datetime.now() - timedelta(days=90)

        historical_dir = os.path.join(self.data_dir, 'historical_data')
        if os.path.exists(historical_dir):
            for root, dirs, files in os.walk(historical_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    try:
                        file_mtime = datetime.fromtimestamp(os.path.getmtime(file_path))
                        if file_mtime < cutoff_date:
                            file_size = os.path.getsize(file_path)
                            os.remove(file_path)
                            files_deleted += 1
                            space_freed += file_size
                    except Exception as e:
                        logger.debug(f"Could not process {file_path}: {e}")

        return files_deleted, space_freed

    def _cleanup_temp_files(self) -> tuple:
        """Clean up temporary files"""
        files_deleted = 0
        space_freed = 0

        temp_dirs = [
            os.path.join(self.data_dir, 'temp'),
            os.path.join(self.data_dir, 'cache'),
            os.path.join(self.data_dir, 'tmp'),
        ]

        for temp_dir in temp_dirs:
            if os.path.exists(temp_dir):
                for root, dirs, files in os.walk(temp_dir):
                    for file in files:
                        file_path = os.path.join(root, file)
                        try:
                            file_size = os.path.getsize(file_path)
                            os.remove(file_path)
                            files_deleted += 1
                            space_freed += file_size
                        except Exception as e:
                            logger.debug(f"Could not delete {file_path}: {e}")

        return files_deleted, space_freed

    def _cleanup_old_logs(self) -> tuple:
        """Clean up log files older than 30 days"""
        files_deleted = 0
        space_freed = 0
        cutoff_date = datetime.now() - timedelta(days=30)

        log_dirs = [
            os.path.join(self.data_dir, 'logs'),
            str(CONFIG.LOGS_DIR) if CONFIG else './logs',
        ]

        for log_dir in log_dirs:
            if os.path.exists(log_dir):
                for file in os.listdir(log_dir):
                    if file.endswith('.log') or file.endswith('.txt'):
                        file_path = os.path.join(log_dir, file)
                        try:
                            file_mtime = datetime.fromtimestamp(os.path.getmtime(file_path))
                            if file_mtime < cutoff_date:
                                file_size = os.path.getsize(file_path)
                                os.remove(file_path)
                                files_deleted += 1
                                space_freed += file_size
                        except Exception as e:
                            logger.debug(f"Could not process {file_path}: {e}")

        return files_deleted, space_freed

    def _cleanup_ai_cache(self) -> tuple:
        """Clean up AI model cache"""
        files_deleted = 0
        space_freed = 0

        cache_dirs = [
            os.path.join(self.data_dir, 'ai_cache'),
            os.path.join(self.data_dir, 'model_cache'),
        ]

        for cache_dir in cache_dirs:
            if os.path.exists(cache_dir):
                for root, dirs, files in os.walk(cache_dir):
                    for file in files:
                        file_path = os.path.join(root, file)
                        try:
                            file_size = os.path.getsize(file_path)
                            os.remove(file_path)
                            files_deleted += 1
                            space_freed += file_size
                        except Exception as e:
                            logger.debug(f"Could not delete {file_path}: {e}")

        return files_deleted, space_freed

    def _cleanup_orphaned_files(self) -> tuple:
        """Clean up orphaned files (files not linked to any profile)"""
        files_deleted = 0
        space_freed = 0

        # Clean empty directories
        for root, dirs, files in os.walk(self.data_dir, topdown=False):
            for dir_name in dirs:
                dir_path = os.path.join(root, dir_name)
                try:
                    if not os.listdir(dir_path):  # Empty directory
                        os.rmdir(dir_path)
                        files_deleted += 1
                except Exception as e:
                    logger.debug(f"Could not remove directory {dir_path}: {e}")

        # Clean .tmp and .bak files
        for root, dirs, files in os.walk(self.data_dir):
            for file in files:
                if file.endswith('.tmp') or file.endswith('.bak') or file.startswith('~'):
                    file_path = os.path.join(root, file)
                    try:
                        file_size = os.path.getsize(file_path)
                        os.remove(file_path)
                        files_deleted += 1
                        space_freed += file_size
                    except Exception as e:
                        logger.debug(f"Could not delete {file_path}: {e}")

        return files_deleted, space_freed

    def _load_backup_history(self):
        """Load backup history from backup directory"""
        try:
            self.backup_table.setRowCount(0)

            # Get backups directory
            backup_dir = str(CONFIG.BACKUPS_DIR) if CONFIG else "./backups"

            if not os.path.exists(backup_dir):
                return

            # Look for backup files in subdirectories
            backup_types = ['daily', 'weekly', 'monthly']
            backups = []

            for backup_type in backup_types:
                type_dir = os.path.join(backup_dir, backup_type)
                if os.path.exists(type_dir):
                    for file in os.listdir(type_dir):
                        if file.endswith('.zip'):
                            file_path = os.path.join(type_dir, file)
                            try:
                                file_stat = os.stat(file_path)
                                backups.append({
                                    'path': file_path,
                                    'type': backup_type.capitalize(),
                                    'date': datetime.fromtimestamp(file_stat.st_mtime),
                                    'size': file_stat.st_size,
                                    'filename': file
                                })
                            except Exception as e:
                                logger.debug(f"Could not read backup file {file_path}: {e}")

            # Also check root backups directory
            for file in os.listdir(backup_dir):
                if file.endswith('.zip'):
                    file_path = os.path.join(backup_dir, file)
                    try:
                        file_stat = os.stat(file_path)
                        backups.append({
                            'path': file_path,
                            'type': 'Manual',
                            'date': datetime.fromtimestamp(file_stat.st_mtime),
                            'size': file_stat.st_size,
                            'filename': file
                        })
                    except Exception as e:
                        logger.debug(f"Could not read backup file {file_path}: {e}")

            # Sort by date (newest first)
            backups.sort(key=lambda x: x['date'], reverse=True)

            # Populate table
            for backup in backups[:50]:  # Show max 50 backups
                row = self.backup_table.rowCount()
                self.backup_table.insertRow(row)

                # Date
                date_item = QTableWidgetItem(backup['date'].strftime('%Y-%m-%d %H:%M'))
                self.backup_table.setItem(row, 0, date_item)

                # Type
                type_item = QTableWidgetItem(backup['type'])
                self.backup_table.setItem(row, 1, type_item)

                # Profile (extract from filename if possible)
                profile = "All Profiles"
                if "profile_" in backup['filename']:
                    try:
                        profile = backup['filename'].split("profile_")[1].split("_")[0]
                    except:
                        pass
                profile_item = QTableWidgetItem(profile)
                self.backup_table.setItem(row, 2, profile_item)

                # Size
                size_mb = backup['size'] / (1024 * 1024)
                size_item = QTableWidgetItem(f"{size_mb:.2f} MB")
                self.backup_table.setItem(row, 3, size_item)

                # Actions (store path for later use)
                action_item = QTableWidgetItem(backup['path'])
                self.backup_table.setItem(row, 4, action_item)

            self.backup_table.resizeColumnsToContents()

        except Exception as e:
            logger.error(f"Error loading backup history: {e}")

    def _gather_obd_data(self, start_date: Optional[datetime], end_date: datetime) -> List[Dict[str, Any]]:
        """Gather OBD data for export"""
        data = []

        try:
            if self.historical_data_manager:
                # Try to get data from historical data manager
                if hasattr(self.historical_data_manager, 'get_readings'):
                    readings = self.historical_data_manager.get_readings(
                        start_date=start_date,
                        end_date=end_date
                    )
                    if readings:
                        data.extend(readings)
                elif hasattr(self.historical_data_manager, 'get_all_readings'):
                    all_readings = self.historical_data_manager.get_all_readings()
                    for reading in all_readings:
                        reading_date = reading.get('timestamp')
                        if reading_date:
                            if isinstance(reading_date, str):
                                reading_date = datetime.fromisoformat(reading_date)
                            if start_date is None or reading_date >= start_date:
                                if reading_date <= end_date:
                                    data.append(reading)
                        else:
                            data.append(reading)

            # Also check for OBD data files
            obd_dir = os.path.join(self.data_dir, 'obd_data')
            if os.path.exists(obd_dir):
                for file in os.listdir(obd_dir):
                    if file.endswith('.json'):
                        file_path = os.path.join(obd_dir, file)
                        try:
                            with open(file_path, 'r', encoding='utf-8') as f:
                                file_data = json.load(f)
                                if isinstance(file_data, list):
                                    data.extend(file_data)
                                elif isinstance(file_data, dict):
                                    data.append(file_data)
                        except Exception as e:
                            logger.debug(f"Could not read OBD file {file_path}: {e}")

        except Exception as e:
            logger.error(f"Error gathering OBD data: {e}")

        return data

    def _gather_service_history(self) -> List[Dict[str, Any]]:
        """Gather service history for export"""
        data = []

        try:
            # Check for service history files
            service_dir = os.path.join(self.data_dir, 'service_history')
            if os.path.exists(service_dir):
                for file in os.listdir(service_dir):
                    if file.endswith('.json'):
                        file_path = os.path.join(service_dir, file)
                        try:
                            with open(file_path, 'r', encoding='utf-8') as f:
                                file_data = json.load(f)
                                if isinstance(file_data, list):
                                    data.extend(file_data)
                                elif isinstance(file_data, dict):
                                    data.append(file_data)
                        except Exception as e:
                            logger.debug(f"Could not read service file {file_path}: {e}")

            # Try to get from vehicle manager if available
            if self.vehicle_manager:
                if hasattr(self.vehicle_manager, 'get_service_history'):
                    service_records = self.vehicle_manager.get_service_history()
                    if service_records:
                        data.extend(service_records)

        except Exception as e:
            logger.error(f"Error gathering service history: {e}")

        return data

    def _gather_predictions(self) -> List[Dict[str, Any]]:
        """Gather AI predictions for export"""
        data = []

        try:
            # Check for prediction files
            predictions_dir = os.path.join(self.data_dir, 'predictions')
            if os.path.exists(predictions_dir):
                for file in os.listdir(predictions_dir):
                    if file.endswith('.json'):
                        file_path = os.path.join(predictions_dir, file)
                        try:
                            with open(file_path, 'r', encoding='utf-8') as f:
                                file_data = json.load(f)
                                if isinstance(file_data, list):
                                    data.extend(file_data)
                                elif isinstance(file_data, dict):
                                    data.append(file_data)
                        except Exception as e:
                            logger.debug(f"Could not read predictions file {file_path}: {e}")

            # Also check AI models directory
            ai_dir = os.path.join(self.data_dir, 'ai_models')
            if os.path.exists(ai_dir):
                for file in os.listdir(ai_dir):
                    if file.endswith('_predictions.json'):
                        file_path = os.path.join(ai_dir, file)
                        try:
                            with open(file_path, 'r', encoding='utf-8') as f:
                                file_data = json.load(f)
                                if isinstance(file_data, list):
                                    data.extend(file_data)
                                elif isinstance(file_data, dict):
                                    data.append(file_data)
                        except Exception as e:
                            logger.debug(f"Could not read AI predictions file {file_path}: {e}")

        except Exception as e:
            logger.error(f"Error gathering predictions: {e}")

        return data

    def _gather_all_data(self, include_obd: bool, include_service: bool, include_predictions: bool) -> Dict[str, Any]:
        """Gather all data types for export"""
        data = {
            'export_date': datetime.now().isoformat(),
            'version': '1.0',
            'data': {}
        }

        try:
            if include_obd:
                obd_data = self._gather_obd_data(None, datetime.now())
                if obd_data:
                    data['data']['obd_readings'] = obd_data

            if include_service:
                service_data = self._gather_service_history()
                if service_data:
                    data['data']['service_history'] = service_data

            if include_predictions:
                predictions_data = self._gather_predictions()
                if predictions_data:
                    data['data']['ai_predictions'] = predictions_data

        except Exception as e:
            logger.error(f"Error gathering all data: {e}")

        return data

    def _write_csv(self, file_path: str, data: List[Dict[str, Any]]):
        """Write data to CSV file"""
        try:
            if not data:
                return

            # Get all unique keys from data
            all_keys = set()
            for item in data:
                if isinstance(item, dict):
                    all_keys.update(item.keys())

            # Convert to sorted list for consistent column order
            fieldnames = sorted(list(all_keys))

            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
                writer.writeheader()

                for item in data:
                    if isinstance(item, dict):
                        # Convert non-string values to strings
                        row = {}
                        for key, value in item.items():
                            if isinstance(value, (dict, list)):
                                row[key] = json.dumps(value)
                            else:
                                row[key] = str(value) if value is not None else ''
                        writer.writerow(row)

        except Exception as e:
            logger.error(f"Error writing CSV file: {e}")
            raise
