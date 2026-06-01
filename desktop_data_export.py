"""
PREDICT - Vehicle Intelligence Platform
Copyright © 2026 PREDICT
All rights reserved.

This file is proprietary and confidential.
Unauthorized copying, modification, distribution, or use is strictly prohibited.

Created: January 2026
Module: Desktop Data Export

Data Export System
Export vehicle data to Excel and CSV formats
- OBD data export
- Trip history export
- Fuel tracking export
- Service history export
- Driving scores export
- Multi-format support (CSV, Excel)

NOTE: This file is named desktop_data_export.py to avoid naming conflict
with the server's data_export.py when server path is in sys.path.
"""

import os
import csv
import logging
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional
import json

from config import get_config
CONFIG = get_config()

logger = logging.getLogger(__name__)


class DataExportSystem:
    """
    Comprehensive data export to CSV and Excel formats

    Export Types:
    - OBD/Historical Data
    - Trip History
    - Fuel Tracking
    - Service History
    - Driving Scores
    - AI Predictions
    - Maintenance Reminders

    Formats:
    - CSV (universal compatibility)
    - Excel (if openpyxl available)
    """

    def __init__(self, historical_data_manager, vehicle_manager,
                 fuel_tracking=None, trip_analytics=None):
        """
        Initialize data export system

        Args:
            historical_data_manager: HistoricalDataManager instance
            vehicle_manager: VehicleProfileManager instance
            fuel_tracking: FuelTrackingSystem instance (optional)
            trip_analytics: TripAnalytics instance (optional)
        """
        self.historical_data = historical_data_manager
        self.vehicle_manager = vehicle_manager
        self.fuel_tracking = fuel_tracking
        self.trip_analytics = trip_analytics

        # Check if openpyxl is available for Excel export
        try:
            import openpyxl
            self.excel_available = True
        except ImportError:
            self.excel_available = False
            logger.warning("openpyxl not available - Excel export disabled")

        # Export directory
        self.export_path = str(CONFIG.DATA_DIR / "exports")
        os.makedirs(self.export_path, exist_ok=True)

        logger.info("Data Export System initialized")

    def export_obd_data(self, profile_id: int, profile_name: str,
                       start_date: Optional[datetime] = None,
                       end_date: Optional[datetime] = None,
                       format: str = 'csv') -> Dict[str, Any]:
        """
        Export OBD/historical data

        Args:
            profile_id: Profile ID
            profile_name: Profile name
            start_date: Start date (None = all data)
            end_date: End date (None = now)
            format: Export format ('csv' or 'excel')

        Returns:
            Export result with file path
        """
        try:
            # Get historical data
            data = self.historical_data.read_profile_data(
                profile_name, profile_id,
                start_date=start_date,
                end_date=end_date
            )

            if not data:
                return {'success': False, 'error': 'No data available'}

            # Generate filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            date_range = ''
            if start_date:
                date_range = f"_{start_date.strftime('%Y%m%d')}_to_{end_date.strftime('%Y%m%d') if end_date else 'now'}"

            filename = f"obd_data_{profile_name}_{profile_id}{date_range}_{timestamp}"

            # Export based on format
            if format == 'excel' and self.excel_available:
                file_path = self._export_obd_to_excel(data, filename, profile_name)
            else:
                file_path = self._export_obd_to_csv(data, filename, profile_name)

            logger.info(f"OBD data exported: {file_path}")

            return {
                'success': True,
                'file_path': file_path,
                'format': format,
                'records': len(data)
            }

        except Exception as e:
            logger.error(f"Error exporting OBD data: {e}")
            return {'success': False, 'error': str(e)}

    def _export_obd_to_csv(self, data: List[Dict], filename: str, profile_name: str) -> str:
        """Export OBD data to CSV"""
        try:
            file_path = os.path.join(self.export_path, f"{filename}.csv")

            # Get all unique keys from data
            all_keys = set()
            for record in data:
                all_keys.update(record.keys())

            # Define column order (important columns first)
            priority_columns = ['timestamp', 'rpm', 'speed', 'speed_kmh', 'coolant_temp',
                              'throttle_position', 'fuel_level_pct', 'engine_load']

            # Sort columns: priority first, then alphabetically
            columns = [col for col in priority_columns if col in all_keys]
            remaining = sorted([col for col in all_keys if col not in priority_columns])
            columns.extend(remaining)

            # Write CSV
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=columns, extrasaction='ignore')

                # Header
                writer.writeheader()

                # Data rows
                for record in data:
                    # Flatten nested dicts (like gps)
                    flat_record = self._flatten_dict(record)
                    writer.writerow(flat_record)

            return file_path

        except Exception as e:
            logger.error(f"Error exporting to CSV: {e}")
            raise

    def _export_obd_to_excel(self, data: List[Dict], filename: str, profile_name: str) -> str:
        """Export OBD data to Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            file_path = os.path.join(self.export_path, f"{filename}.xlsx")

            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "OBD Data"

            # Get columns
            all_keys = set()
            for record in data:
                all_keys.update(record.keys())

            priority_columns = ['timestamp', 'rpm', 'speed', 'speed_kmh', 'coolant_temp',
                              'throttle_position', 'fuel_level_pct', 'engine_load']

            columns = [col for col in priority_columns if col in all_keys]
            remaining = sorted([col for col in all_keys if col not in priority_columns])
            columns.extend(remaining)

            # Write header with styling
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for col_idx, column in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=column)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # Write data rows
            for row_idx, record in enumerate(data, 2):
                flat_record = self._flatten_dict(record)
                for col_idx, column in enumerate(columns, 1):
                    value = flat_record.get(column, '')
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save
            wb.save(file_path)

            return file_path

        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            raise

    def export_trip_history(self, profile_id: int, profile_name: str,
                           days: int = 30, format: str = 'csv') -> Dict[str, Any]:
        """
        Export trip history

        Args:
            profile_id: Profile ID
            profile_name: Profile name
            days: Days of history
            format: Export format

        Returns:
            Export result
        """
        try:
            if not self.trip_analytics:
                return {'success': False, 'error': 'Trip analytics not available'}

            # Load trips
            trips = self.trip_analytics._load_trips(profile_name, profile_id, days)

            if not trips:
                return {'success': False, 'error': 'No trip data available'}

            # Generate filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"trips_{profile_name}_{profile_id}_{days}days_{timestamp}"

            # Flatten trip data
            flat_trips = []
            for trip in trips:
                flat_trip = {
                    'trip_id': trip.get('trip_id'),
                    'start_time': trip.get('start_time'),
                    'end_time': trip.get('end_time'),
                    'duration_minutes': round(trip.get('duration_seconds', 0) / 60, 1),
                    'distance_km': round(trip.get('total_distance', 0), 2),
                    'avg_speed_kmh': round(trip.get('avg_speed', 0), 1),
                    'max_speed_kmh': round(trip.get('max_speed', 0), 1),
                    'fuel_consumed_liters': trip.get('fuel_consumed'),
                    'fuel_efficiency_l_per_100km': trip.get('fuel_efficiency_l_per_100km'),
                    'fuel_efficiency_mpg': trip.get('fuel_efficiency_mpg'),
                    'start_lat': trip.get('start_location', {}).get('latitude'),
                    'start_lon': trip.get('start_location', {}).get('longitude'),
                    'end_lat': trip.get('end_location', {}).get('latitude'),
                    'end_lon': trip.get('end_location', {}).get('longitude'),
                    'route_points_count': len(trip.get('route_points', []))
                }
                flat_trips.append(flat_trip)

            # Export
            if format == 'excel' and self.excel_available:
                file_path = self._export_to_excel_simple(flat_trips, filename, "Trip History")
            else:
                file_path = self._export_to_csv_simple(flat_trips, filename)

            logger.info(f"Trip history exported: {file_path}")

            return {
                'success': True,
                'file_path': file_path,
                'format': format,
                'records': len(flat_trips)
            }

        except Exception as e:
            logger.error(f"Error exporting trip history: {e}")
            return {'success': False, 'error': str(e)}

    def export_fuel_tracking(self, profile_id: int, days: int = 90,
                            format: str = 'csv') -> Dict[str, Any]:
        """Export fuel tracking data"""
        try:
            if not self.fuel_tracking:
                return {'success': False, 'error': 'Fuel tracking not available'}

            # Get fillup history
            fillups = self.fuel_tracking.get_fillup_history(profile_id, days)

            if not fillups:
                return {'success': False, 'error': 'No fillup data available'}

            # Generate filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"fuel_tracking_{profile_id}_{days}days_{timestamp}"

            # Flatten fillup data
            flat_fillups = []
            for fillup in fillups:
                flat_fillup = {
                    'fillup_id': fillup.get('fillup_id'),
                    'timestamp': fillup.get('timestamp'),
                    'liters': fillup.get('liters'),
                    'gallons': fillup.get('gallons'),
                    'cost': fillup.get('cost'),
                    'cost_per_liter': fillup.get('cost_per_liter'),
                    'odometer_km': fillup.get('odometer_km'),
                    'odometer_miles': fillup.get('odometer_miles'),
                    'full_tank': fillup.get('full_tank'),
                    'fuel_grade': fillup.get('fuel_grade'),
                    'station_name': fillup.get('station_name')
                }

                # Add calculated metrics if available
                metrics = fillup.get('calculated_metrics', {})
                if metrics:
                    flat_fillup.update({
                        'distance_since_last_km': metrics.get('distance_since_last_km'),
                        'fuel_consumed_liters': metrics.get('fuel_consumed_liters'),
                        'fuel_efficiency_mpg': metrics.get('fuel_efficiency_mpg'),
                        'fuel_efficiency_l_per_100km': metrics.get('fuel_efficiency_l_per_100km'),
                        'cost_per_km': metrics.get('cost_per_km')
                    })

                flat_fillups.append(flat_fillup)

            # Export
            if format == 'excel' and self.excel_available:
                file_path = self._export_to_excel_simple(flat_fillups, filename, "Fuel Tracking")
            else:
                file_path = self._export_to_csv_simple(flat_fillups, filename)

            logger.info(f"Fuel tracking exported: {file_path}")

            return {
                'success': True,
                'file_path': file_path,
                'format': format,
                'records': len(flat_fillups)
            }

        except Exception as e:
            logger.error(f"Error exporting fuel tracking: {e}")
            return {'success': False, 'error': str(e)}

    def export_all_data(self, profile_id: int, profile_name: str,
                       days: int = 30, format: str = 'csv') -> Dict[str, Any]:
        """
        Export all data types in one go

        Args:
            profile_id: Profile ID
            profile_name: Profile name
            days: Days of data
            format: Export format

        Returns:
            Export result with list of files
        """
        try:
            exported_files = []
            errors = []

            # 1. OBD Data
            obd_result = self.export_obd_data(
                profile_id, profile_name,
                start_date=datetime.now() - timedelta(days=days),
                format=format
            )
            if obd_result.get('success'):
                exported_files.append(obd_result['file_path'])
            else:
                errors.append(f"OBD: {obd_result.get('error')}")

            # 2. Trip History
            if self.trip_analytics:
                trip_result = self.export_trip_history(
                    profile_id, profile_name, days, format
                )
                if trip_result.get('success'):
                    exported_files.append(trip_result['file_path'])
                else:
                    errors.append(f"Trips: {trip_result.get('error')}")

            # 3. Fuel Tracking
            if self.fuel_tracking:
                fuel_result = self.export_fuel_tracking(profile_id, days, format)
                if fuel_result.get('success'):
                    exported_files.append(fuel_result['file_path'])
                else:
                    errors.append(f"Fuel: {fuel_result.get('error')}")

            return {
                'success': len(exported_files) > 0,
                'exported_files': exported_files,
                'file_count': len(exported_files),
                'errors': errors if errors else None
            }

        except Exception as e:
            logger.error(f"Error exporting all data: {e}")
            return {'success': False, 'error': str(e)}

    def _export_to_csv_simple(self, data: List[Dict], filename: str) -> str:
        """Simple CSV export"""
        try:
            file_path = os.path.join(self.export_path, f"{filename}.csv")

            if not data:
                return file_path

            # Get columns from first record
            columns = list(data[0].keys())

            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=columns)
                writer.writeheader()
                writer.writerows(data)

            return file_path

        except Exception as e:
            logger.error(f"Error in simple CSV export: {e}")
            raise

    def _export_to_excel_simple(self, data: List[Dict], filename: str, sheet_name: str) -> str:
        """Simple Excel export"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill

            file_path = os.path.join(self.export_path, f"{filename}.xlsx")

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name

            if not data:
                wb.save(file_path)
                return file_path

            # Get columns
            columns = list(data[0].keys())

            # Write header
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for col_idx, column in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=column)
                cell.fill = header_fill
                cell.font = header_font

            # Write data
            for row_idx, record in enumerate(data, 2):
                for col_idx, column in enumerate(columns, 1):
                    ws.cell(row=row_idx, column=col_idx, value=record.get(column))

            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(file_path)
            return file_path

        except Exception as e:
            logger.error(f"Error in simple Excel export: {e}")
            raise

    def _flatten_dict(self, d: Dict, parent_key: str = '', sep: str = '_') -> Dict:
        """Flatten nested dictionary"""
        items = []
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k

            if isinstance(v, dict):
                items.extend(self._flatten_dict(v, new_key, sep=sep).items())
            elif isinstance(v, list):
                # Convert lists to JSON string
                items.append((new_key, json.dumps(v)))
            else:
                items.append((new_key, v))

        return dict(items)
